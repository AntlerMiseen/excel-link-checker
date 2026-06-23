# -*- coding: utf-8 -*-
"""测试 URL 提取 鈥?extract_urls_from_cell / extract_urls_from_workbook"""

import pytest
from link_checker_core import extract_urls_from_cell, extract_urls_from_workbook, UrlEntry


class TestExtractUrlsFromCell:
    """单元测试：从单个 Cell 值提取 URL"""

    def test_single_url(self):
        assert extract_urls_from_cell("https://example.com") == ["https://example.com"]

    def test_multiple_urls(self):
        text = "访问 https://a.com 或 http://b.com/page"
        result = extract_urls_from_cell(text)
        assert "https://a.com" in result
        assert "http://b.com/page" in result

    def test_no_url(self):
        assert extract_urls_from_cell("纯文本无链接") == []

    def test_none_value(self):
        assert extract_urls_from_cell(None) == []

    def test_nan_value(self):
        assert extract_urls_from_cell(float("nan")) == []

    def test_empty_string(self):
        assert extract_urls_from_cell("") == []

    def test_number_value(self):
        assert extract_urls_from_cell(12345) == []

    def test_http_url(self):
        assert extract_urls_from_cell("http://example.com") == ["http://example.com"]

    def test_url_with_query(self):
        text = "https://example.com/search?q=python&lang=zh"
        assert extract_urls_from_cell(text) == [text]

    def test_url_with_fragment(self):
        text = "https://example.com/page#section"
        assert extract_urls_from_cell(text) == [text]

    def test_url_with_port(self):
        text = "http://localhost:8080/api"
        assert extract_urls_from_cell(text) == ["http://localhost:8080/api"]

    def test_ascii_punctuation_stripped(self):
        """ASCII 标点（逗号、句号等）会被 strip"""
        text = "https://example.com,"
        result = extract_urls_from_cell(text)
        assert result == ["https://example.com"]

    def test_chinese_punctuation_not_stripped(self):
        """中文标点（。）不在 rstrip 字符集中，不会被 strip"""
        text = "https://example.com。"
        result = extract_urls_from_cell(text)
        # 中文句号不会被 strip，原样保留
        assert result == ["https://example.com。"]

    def test_url_surrounded_by_parentheses(self):
        """括号内的链接可以被提取（闭合括号不匹配被 strip）"""
        text = "链接 (https://example.com) 在这里"
        result = extract_urls_from_cell(text)
        # 正则提取 https://example.com) 然后 rstrip 去除 )
        assert "https://example.com" in result

    def test_ftp_not_extracted(self):
        assert extract_urls_from_cell("ftp://files.example.com") == []

    def test_duplicate_urls_preserved(self):
        text = "a https://x.com b https://x.com"
        result = extract_urls_from_cell(text)
        assert result.count("https://x.com") == 2

    def test_mixed_content(self):
        text = "标题：测试\n链接：https://test.com\n描述：something"
        result = extract_urls_from_cell(text)
        assert result == ["https://test.com"]


class TestExtractUrlsFromWorkbook:
    """从 Workbook 提取 URL"""

    def test_extracts_correct_count(self, sample_workbook):
        entries = extract_urls_from_workbook(sample_workbook)
        # Row 2: google(1) + Row 3: baidu(1) + Row 4: 10.0.0.1(1)
        # Row 5: ftp (0) + Row 6: no URL (0) + Row 7: a.com+b.com(2)
        # Row 8: None (0) + Row 9: numbers (0) = 5 total
        assert len(entries) == 5

    def test_entries_are_url_entry_objects(self, sample_workbook):
        entries = extract_urls_from_workbook(sample_workbook)
        for e in entries:
            assert isinstance(e, UrlEntry)

    def test_sheet_name_correct(self, sample_workbook):
        entries = extract_urls_from_workbook(sample_workbook)
        assert all(e.sheet == "Sheet1" for e in entries)

    def test_row_numbers_are_2_based(self, sample_workbook):
        entries = extract_urls_from_workbook(sample_workbook)
        assert all(e.row >= 2 for e in entries)

    def test_multi_link_row(self, sample_workbook):
        entries = extract_urls_from_workbook(sample_workbook)
        # 多链接行是第 7 行（row=7，0-based index=6）
        multi = [e for e in entries if e.row == 7]
        assert len(multi) == 2
        urls = {e.url for e in multi}
        assert "https://a.com" in urls
        assert "https://b.com" in urls

    def test_ftp_not_extracted_in_workbook(self, sample_workbook):
        entries = extract_urls_from_workbook(sample_workbook)
        urls = [e.url for e in entries]
        assert not any("ftp" in u for u in urls)

    def test_empty_sheet_returns_empty(self):
        from openpyxl import Workbook
        wb = Workbook()
        entries = extract_urls_from_workbook(wb)
        assert entries == []


    def test_hyperlink_target_priority(self):
        """Excel 超链接对象优先于文本正则匹配"""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Hyperlink'
        ws.append(['名称', '链接'])
        ws.append(['文本链接', 'https://text.example.com'])
        # 设置超链接对象（模拟 插入→链接 的效果）
        cell = ws.cell(row=2, column=2)
        cell.hyperlink = 'https://hyperlink.example.com'
        cell.value = 'click here'  # 显示文本不是 URL

        entries = extract_urls_from_workbook(wb)
        urls = [e.url for e in entries]
        # 应优先提取 hyperlink.target，而非 cell.value 文本
        assert 'https://hyperlink.example.com' in urls
        assert 'https://text.example.com' not in urls