# -*- coding: utf-8 -*-
"""测试 Excel 处理 鈥?process_excel / _build_summary_sheet / 工具函数"""

import os
import tempfile
import pytest
from unittest.mock import patch, MagicMock
from openpyxl import load_workbook
from link_checker_core import (
    process_excel, _build_summary_sheet,
    generate_output_path, hash_file,
    ProcessResult, Config, cfg,
)


class TestProcessExcel:
    """process_excel 集成测试"""

    @pytest.fixture(autouse=True)
    def mock_batch_check(self):
        with patch("link_checker_core.batch_check_urls") as mock_batch:
            yield mock_batch

    def test_no_urls_in_file(self, mock_batch_check):
        from openpyxl import Workbook
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        wb = Workbook()
        ws = wb.active
        ws.append(["名称", "备注"])
        ws.append(["纯文本", "无链接"])
        wb.save(path)

        try:
            output = path.replace(".xlsx", "_out.xlsx")
            result = process_excel(path, output)
            assert result.total_urls == 0
            assert result.valid_count == 0
            assert result.invalid_count == 0
        finally:
            os.unlink(path)
            if os.path.exists(output):
                os.unlink(output)

    def test_file_链接检测结果_found(self):
        with pytest.raises(FileNotFoundError):
            process_excel("不存在的文件.xlsx", "out.xlsx")

    def test_output_file_created(self, temp_xlsx, mock_batch_check):
        mock_batch_check.return_value = {}
        output = temp_xlsx.replace(".xlsx", "_检测_out.xlsx")
        try:
            result = process_excel(temp_xlsx, output)
            assert os.path.exists(result.output_path)
        finally:
            if os.path.exists(output):
                os.unlink(output)
            import glob
            base = temp_xlsx.replace(".xlsx", "_检测_out")
            for f in glob.glob(base + "*.xlsx"):
                try:
                    os.unlink(f)
                except OSError:
                    pass

    def test_yellow_highlight_on_invalid(self, temp_xlsx, mock_batch_check):
        from link_checker_core import CheckResult

        def mock_result(entries, *args, **kwargs):
            return {
                "https://www.google.com": CheckResult("https://www.google.com", True, "OK", 0),
                "https://www.baidu.com": CheckResult("https://www.baidu.com", False, "超时", 2),
            }

        mock_batch_check.side_effect = mock_result
        output = temp_xlsx.replace(".xlsx", "_yellow_out.xlsx")
        try:
            result = process_excel(temp_xlsx, output)
            assert os.path.exists(result.output_path)
            wb = load_workbook(result.output_path)
            ws = wb["Sheet1"]
            # 百度行 (row 3) 应标黄
            cell = ws.cell(row=3, column=1)
            assert cell.fill.start_color.rgb == "00FFFF00" or \
                   "FFFF00" in str(cell.fill.start_color.rgb)
        finally:
            if os.path.exists(output):
                os.unlink(output)

    def test_summary_sheet_created(self, temp_xlsx, mock_batch_check):
        from link_checker_core import CheckResult
        mock_batch_check.return_value = {
            "https://www.google.com": CheckResult("https://www.google.com", True, "OK", 0),
            "https://www.baidu.com": CheckResult("https://www.baidu.com", False, "超时", 2),
        }
        output = temp_xlsx.replace(".xlsx", "_summary_out.xlsx")
        try:
            result = process_excel(temp_xlsx, output)
            wb = load_workbook(result.output_path)
            assert any("链接检测" in name for name in wb.sheetnames)
        finally:
            if os.path.exists(output):
                os.unlink(output)

    def test_progress_callback_invoked(self, temp_xlsx, mock_batch_check):
        mock_batch_check.return_value = {}
        calls = []

        def cb(completed, total, result):
            calls.append((completed, total))

        output = temp_xlsx.replace(".xlsx", "_cb_out.xlsx")
        try:
            process_excel(temp_xlsx, output, progress_callback=cb)
            assert len(calls) >= 1
        finally:
            if os.path.exists(output):
                os.unlink(output)

    def test_original_file_untouched(self, temp_xlsx, mock_batch_check):
        original_mtime = os.path.getmtime(temp_xlsx)
        mock_batch_check.return_value = {}
        output = temp_xlsx.replace(".xlsx", "_untouched_out.xlsx")
        try:
            process_excel(temp_xlsx, output)
            assert os.path.getmtime(temp_xlsx) == original_mtime
        finally:
            if os.path.exists(output):
                os.unlink(output)

    def test_valid_row_not_highlighted(self, temp_xlsx, mock_batch_check):
        from link_checker_core import CheckResult
        mock_batch_check.return_value = {
            "https://www.google.com": CheckResult("https://www.google.com", True, "OK", 0),
            "https://www.baidu.com": CheckResult("https://www.baidu.com", True, "OK", 0),
        }
        output = temp_xlsx.replace(".xlsx", "_valid_out.xlsx")
        try:
            result = process_excel(temp_xlsx, output)
            wb = load_workbook(result.output_path)
            ws = wb["Sheet1"]
            cell = ws.cell(row=2, column=1)
            rgb = str(cell.fill.start_color.rgb) if cell.fill.start_color else ""
            assert "FFFF00" not in rgb
        finally:
            if os.path.exists(output):
                os.unlink(output)


class TestBuildSummarySheet:
    """_build_summary_sheet 测试"""

    def test_creates_sheet_at_index_0(self):
        from openpyxl import Workbook
        wb = Workbook()
        ws_data = wb.active
        ws_data.title = "Data"
        ws_data.append(["A"])

        result_list = [
            {"sheet": "Data", "row": 2, "col": "链接", "url": "https://ok.com",
             "valid": True, "detail": "HTTP 200", "retries": 0},
        ]
        _build_summary_sheet(wb, result_list, 1, 0, 0)
        assert wb.sheetnames[0] == "链接检测结果"
        ws = wb["链接检测结果"]
        assert ws.cell(row=4, column=1).value == "Sheet"

    def test_overwrites_existing_summary_sheet(self):
        from openpyxl import Workbook
        wb = Workbook()
        old_ws = wb.active
        old_ws.title = "链接检测结果"
        old_ws.append(["旧数据"])
        wb.create_sheet("Data")
        wb["Data"].append(["A"])

        result_list = [
            {"sheet": "Data", "row": 2, "col": "链接", "url": "https://ok.com",
             "valid": True, "detail": "HTTP 200", "retries": 0},
        ]
        _build_summary_sheet(wb, result_list, 1, 0, 0)
        assert wb.sheetnames[0] == "链接检测结果"
        assert wb["链接检测结果"].cell(row=1, column=1).value != "旧数据"

    def test_statistics_row(self):
        from openpyxl import Workbook
        wb = Workbook()
        result_list = [
            {"sheet": "S1", "row": 2, "col": "A", "url": "https://a.com",
             "valid": True, "detail": "OK", "retries": 0},
            {"sheet": "S1", "row": 3, "col": "A", "url": "https://b.com",
             "valid": False, "detail": "超时", "retries": 2},
            {"sheet": "S1", "row": 4, "col": "A", "url": "https://c.com",
             "valid": True, "detail": "OK", "retries": 0},
        ]
        _build_summary_sheet(wb, result_list, 2, 1, 0)
        ws = wb["链接检测结果"]
        stat = ws.cell(row=2, column=1).value
        assert "有效" in stat or "2" in stat

    def test_header_formatting(self):
        from openpyxl import Workbook
        wb = Workbook()
        _build_summary_sheet(wb, [], 0, 0, 0)
        ws = wb["链接检测结果"]
        header_cell = ws.cell(row=4, column=1)
        assert "4472C4" in str(header_cell.fill.start_color.rgb)

    def test_valid_invalid_coloring(self):
        from openpyxl import Workbook
        wb = Workbook()
        result_list = [
            {"sheet": "S", "row": 2, "col": "A", "url": "https://ok.com",
             "valid": True, "detail": "OK", "retries": 0},
            {"sheet": "S", "row": 3, "col": "A", "url": "https://bad.com",
             "valid": False, "detail": "超时", "retries": 1},
        ]
        _build_summary_sheet(wb, result_list, 1, 1, 0)
        ws = wb["链接检测结果"]
        valid_cell = ws.cell(row=5, column=5)
        assert "C6EFCE" in str(valid_cell.fill.start_color.rgb)
        invalid_cell = ws.cell(row=6, column=5)
        assert "FFC7CE" in str(invalid_cell.fill.start_color.rgb)


class TestGenerateOutputPath:
    """输出路径生成测试"""

    def test_basic_pattern(self):
        path = generate_output_path("test.xlsx")
        assert "result" in path and "test_链接检测结果_" in path
        assert path.endswith(".xlsx")

    def test_timestamp_present(self):
        import re
        path = generate_output_path("data.xlsx")
        assert re.search(r"_\d{8}_\d{6}_\d{6}\.xlsx$", path)

    def test_subfolder_created(self, tmp_path):
        d = str(tmp_path)
        input_file = os.path.join(d, "test.xlsx")
        from openpyxl import Workbook
        wb = Workbook()
        wb.save(input_file)
        path = generate_output_path(input_file)
        assert os.path.isdir(os.path.join(d, "result"))
        assert path.startswith(os.path.join(d, "result"))

    def test_different_inputs_different_names(self):
        p1 = generate_output_path("a.xlsx")
        p2 = generate_output_path("b.xlsx")
        assert os.path.basename(p1) != os.path.basename(p2)


class TestHashFile:
    """文件哈希测试"""

    def test_sha256(self, temp_xlsx):
        h = hash_file(temp_xlsx)
        assert len(h) == 64
        assert all(c in "0123456789abcdef" for c in h)

    def test_sha256_deterministic(self, temp_xlsx):
        h1 = hash_file(temp_xlsx)
        h2 = hash_file(temp_xlsx)
        assert h1 == h2

    def test_md5_algorithm(self, temp_xlsx):
        h = hash_file(temp_xlsx, algorithm="md5")
        assert len(h) == 32
