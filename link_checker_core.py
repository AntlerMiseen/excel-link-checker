#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel链接有效性检测 - 企业级核心模块
提供日志、配置、URL安全检查、重试机制、Session复用、格式保留等功能。
"""

import re
import requests
import socket
import os
import sys
import io
import time
import hashlib
import ipaddress
import logging
from concurrent.futures import ThreadPoolExecutor
import concurrent.futures
import threading
from datetime import datetime
from dataclasses import dataclass, field
from typing import Callable, Optional
from urllib.parse import urlparse


import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter


# ==================== 配置 ====================
@dataclass(frozen=True)
class Config:
    """全局配置，集中管理所有可调参数"""
    TIMEOUT: int = 10
    MAX_WORKERS: int = 10
    MAX_RETRIES: int = 2
    RETRY_BACKOFF: float = 1.5
    MAX_REDIRECTS: int = 5
    BATCH_SIZE: int = 200
    LARGE_BATCH_THRESHOLD: int = 500

    USER_AGENT: str = (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/125.0.0.0 Safari/537.36"
    )

    YELLOW_HEX: str = "FFFF00"
    HEADER_BG_HEX: str = "4472C4"

    LOCAL_NETWORKS: tuple = (
        ipaddress.ip_network("127.0.0.0/8"),
        ipaddress.ip_network("10.0.0.0/8"),
        ipaddress.ip_network("172.16.0.0/12"),
        ipaddress.ip_network("192.168.0.0/16"),
        ipaddress.ip_network("169.254.0.0/16"),
        ipaddress.ip_network("::1/128"),
        ipaddress.ip_network("fc00::/7"),
        ipaddress.ip_network("fe80::/10"),
    )

    BLOCKED_SCHEMES: tuple = ("file", "ftp", "data", "javascript", "vbscript")
    BLOCKED_TLDS: tuple = (".local", ".internal", ".corp", ".home", ".lan")

    LOG_FILE: str = "link_checker.log"
    LOG_MAX_BYTES: int = 5 * 1024 * 1024
    LOG_BACKUP_COUNT: int = 3


cfg = Config()
# 全局 socket 超时，防止 DNS 解析无限阻塞
socket.setdefaulttimeout(cfg.TIMEOUT)

# ==================== 日志系统 ====================
_logger: Optional[logging.Logger] = None


def get_logger() -> logging.Logger:
    global _logger
    if _logger is not None:
        return _logger

    _logger = logging.getLogger("LinkChecker")
    _logger.setLevel(logging.DEBUG)

    # 控制台 handler（INFO 以上）
    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", "%H:%M:%S"))
    _logger.addHandler(ch)

    # 文件 handler（DEBUG 以上，自动轮转）
    try:
        from logging.handlers import RotatingFileHandler
        fh = RotatingFileHandler(
            cfg.LOG_FILE, maxBytes=cfg.LOG_MAX_BYTES,
            backupCount=cfg.LOG_BACKUP_COUNT, encoding="utf-8"
        )
        fh.setLevel(logging.DEBUG)
        fh.setFormatter(logging.Formatter(
            "%(asctime)s [%(levelname)s] %(filename)s:%(lineno)d %(message)s"
        ))
        _logger.addHandler(fh)
    except PermissionError:
        pass

    return _logger


log = get_logger()


# ==================== 数据结构 ====================
@dataclass
class UrlEntry:
    sheet: str
    row: int        # 1-based Excel row
    col_idx: int    # 0-based column index
    col_name: str
    url: str


@dataclass
class CheckResult:
    url: str
    valid: bool
    detail: str
    retries: int = 0


@dataclass
class ProcessResult:
    input_path: str
    output_path: str
    total_urls: int
    valid_count: int
    invalid_count: int
    skipped_count: int
    results: list = field(default_factory=list)
    errors: list = field(default_factory=list)
    light_output_path: str = ""


# ==================== 样式常量 ====================
FILL_YELLOW = PatternFill(start_color=cfg.YELLOW_HEX, end_color=cfg.YELLOW_HEX, fill_type="solid")
FILL_VALID = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_INVALID = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_HEADER = PatternFill(start_color=cfg.HEADER_BG_HEX, end_color=cfg.HEADER_BG_HEX, fill_type="solid")
FONT_HEADER = Font(color="FFFFFF", bold=True, size=11)
FONT_VALID = Font(color="006100")
FONT_INVALID = Font(color="9C0006")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

URL_PATTERN = re.compile(r"https?://[^\s'\"<>]+", re.IGNORECASE)


# ==================== URL 安全校验 ====================
def is_url_safe(url: str) -> tuple[bool, str]:
    """检查URL是否安全可请求，返回(安全, 原因)"""
    try:
        parsed = urlparse(url)
    except Exception:
        return False, "URL解析失败"

    # 协议检查
    scheme = (parsed.scheme or "").lower()
    if scheme not in ("http", "https"):
        return False, f"不允许的协议: {scheme or '无'}"

    hostname = (parsed.hostname or "").lower()

    # 空主机名
    if not hostname:
        return False, "缺少主机名"

    # TLD 黑名单
    for tld in cfg.BLOCKED_TLDS:
        if hostname.endswith(tld):
            return False, f"禁止的内部域名后缀: {tld}"

    # IP 检查（localhost + 内网）
    try:
        ip = ipaddress.ip_address(hostname)
        if ip.is_loopback:
            return False, "禁止回环地址 (localhost)"
        if ip.is_link_local:
            return False, "禁止链路本地地址"
        for net in cfg.LOCAL_NETWORKS:
            if ip in net:
                return False, f"禁止内网地址: {net}"
    except ValueError:
        pass  # 非IP地址，正常域名，继续

    return True, "OK"


# ==================== URL 提取 ====================
def extract_urls_from_cell(value) -> list[str]:
    """从单个单元格值中提取URL"""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return []
    text = str(value)
    urls = URL_PATTERN.findall(text)
    return [u.rstrip(".,;:!?)〕】」』\"'") for u in urls]


def extract_urls_from_workbook(wb) -> list[UrlEntry]:
    """遍历所有Sheet，提取URL及其位置（优先读取超链接对象，再回退正则匹配文本）"""
    entries: list[UrlEntry] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=1))
        if not rows or len(rows) < 2:
            continue
        # 第一行是表头，提取列名
        headers = [cell.value for cell in rows[0]]
        for row_idx, row_cells in enumerate(rows[1:], start=2):
            for col_idx, cell in enumerate(row_cells):
                urls = []
                # 优先读取 Excel 超链接对象（插入→链接 设置的）
                if cell.hyperlink and cell.hyperlink.target:
                    target = str(cell.hyperlink.target)
                    if target.startswith(('http://', 'https://')):
                        urls.append(target)
                # 回退到正则匹配纯文本 URL
                if not urls:
                    urls = extract_urls_from_cell(cell.value)
                col_name = headers[col_idx] if col_idx < len(headers) else f"列{col_idx+1}"
                for url in urls:
                    entries.append(UrlEntry(
                        sheet=sheet_name, row=row_idx,
                        col_idx=col_idx, col_name=col_name, url=url
                    ))
    return entries


# ==================== 网络检测 ====================
# 线程局部存储：每个线程独立 Session，避免并发死锁
_thread_local = threading.local()


def _get_session() -> requests.Session:
    """获取当前线程的独立 Session（线程安全），无内部重试，由外层统一控制"""
    if not hasattr(_thread_local, "session"):
        session = requests.Session()
        session.headers.update({"User-Agent": cfg.USER_AGENT})
        session.max_redirects = cfg.MAX_REDIRECTS
        _thread_local.session = session
    return _thread_local.session


def check_single_url(url: str) -> CheckResult:
    """检测单个URL，自动重试"""
    retries = 0
    last_error = ""

    for attempt in range(cfg.MAX_RETRIES + 1):
        try:
            session = _get_session()
            resp = session.head(url, timeout=(cfg.TIMEOUT, cfg.TIMEOUT), allow_redirects=True)
            if resp.status_code < 400:
                return CheckResult(url=url, valid=True, detail=f"HTTP {resp.status_code}", retries=retries)
            resp = session.get(url, timeout=(cfg.TIMEOUT, cfg.TIMEOUT), allow_redirects=True, stream=True)
            resp.close()
            if resp.status_code < 400:
                return CheckResult(url=url, valid=True, detail=f"HTTP {resp.status_code}", retries=retries)
            return CheckResult(url=url, valid=False, detail=f"HTTP {resp.status_code}", retries=retries)

        except requests.exceptions.Timeout:
            last_error = "连接超时"
        except requests.exceptions.SSLError:
            last_error = "SSL证书错误"
            break  # 不重试
        except requests.exceptions.TooManyRedirects:
            last_error = "重定向过多"
            break  # 不重试
        except requests.exceptions.ConnectionError as e:
            last_error = "无法连接"
            # DNS解析失败等不值得重试
            if "NameResolutionError" in type(e).__name__ or "getaddrinfo" in str(e).lower():
                return CheckResult(url=url, valid=False, detail="DNS解析失败", retries=retries)
        except requests.exceptions.RequestException as e:
            last_error = f"请求异常: {type(e).__name__}"

        if attempt < cfg.MAX_RETRIES:
            retries += 1
            delay = cfg.RETRY_BACKOFF ** attempt
            log.debug(f"重试 {retries}/{cfg.MAX_RETRIES}: {url} ({last_error})，等待 {delay:.1f}s")
            time.sleep(delay)

    return CheckResult(url=url, valid=False, detail=last_error, retries=retries)


def batch_check_urls(
    entries: list[UrlEntry],
    progress_callback: Optional[Callable] = None
) -> dict[str, CheckResult]:
    """并发检测URL，支持分批和进度回调"""
    if not entries:
        return {}

    # 去重
    unique_urls = list(dict.fromkeys(e.url for e in entries))
    results: dict[str, CheckResult] = {}

    # URL 安全检查（预过滤）
    safe_urls = []
    for url in unique_urls:
        ok, reason = is_url_safe(url)
        if ok:
            safe_urls.append(url)
        else:
            results[url] = CheckResult(url=url, valid=False, detail=f"安全拦截: {reason}")
            log.warning(f"URL安全拦截: {url} — {reason}")

    if not safe_urls:
        return results

    # ????
    batches = [safe_urls[i:i + cfg.BATCH_SIZE]
               for i in range(0, len(safe_urls), cfg.BATCH_SIZE)]

    total = len(safe_urls)
    completed = 0

    for batch in batches:
        executor = ThreadPoolExecutor(max_workers=cfg.MAX_WORKERS)
        timed_out = False
        try:
            future_map = {executor.submit(check_single_url, u): u for u in batch}
            pending = set(future_map.keys())
            deadline = cfg.TIMEOUT * (cfg.MAX_RETRIES + 2) + 15
            while pending:
                done, pending = concurrent.futures.wait(
                    pending, timeout=deadline,
                    return_when=concurrent.futures.FIRST_COMPLETED
                )
                if not done and pending:
                    timed_out = True
                    for f in pending:
                        f.cancel()
                        url = future_map[f]
                        results[url] = CheckResult(url=url, valid=False, detail="检测超时")
                        completed += 1
                        if progress_callback:
                            progress_callback(completed, total,
                                CheckResult(url=url, valid=False, detail="检测超时"))
                    break
                for f in done:
                    try:
                        result = f.result(timeout=5)
                    except Exception as e:
                        url = future_map[f]
                        result = CheckResult(url=url, valid=False, detail=f"内部错误: {e}")
                        log.debug(f"future异常: {url} — {e}")
                    results[result.url] = result
                    completed += 1
                    if progress_callback:
                        progress_callback(completed, total, result)
        finally:
            # shutdown(wait=False) 不等待卡死的线程，避免永久阻塞
            executor.shutdown(wait=not timed_out)

    return results

# ==================== Excel 处理 ====================
def process_excel(
    input_path: str,
    output_path: str,
    progress_callback: Optional[Callable] = None,
    output_full: bool = True,
    output_light: bool = True
) -> ProcessResult:
    """检测链接并输出结果。output_full: 完整版(整行标黄+汇总Sheet), output_light: 仅标黄版(整行标黄,无新增Sheet)"""
    log.info(f"读取文件: {input_path}")

    if not os.path.exists(input_path):
        raise FileNotFoundError(f"文件不存在: {input_path}")

    wb = load_workbook(input_path)
    entries = extract_urls_from_workbook(wb)
    total_urls = len(entries)

    if total_urls == 0:
        log.warning("未发现任何链接")
        wb.close()
        return ProcessResult(
            input_path=input_path, output_path=output_path,
            total_urls=0, valid_count=0, invalid_count=0, skipped_count=0
        )

    log.info(f"发现 {total_urls} 个链接，开始检测...")
    if progress_callback:
        progress_callback(0, total_urls, None)

    check_results = batch_check_urls(entries, progress_callback)

    # 统计
    valid_count = 0
    invalid_count = 0
    skipped_count = 0
    result_list = []
    invalid_keys: set[tuple[str, int]] = set()

    for entry in entries:
        cr = check_results.get(entry.url)
        if cr is None:
            skipped_count += 1
            continue
        if cr.valid:
            valid_count += 1
        else:
            invalid_count += 1
            invalid_keys.add((entry.sheet, entry.row))

        result_list.append({
            "sheet": entry.sheet, "row": entry.row,
            "col": entry.col_name, "url": entry.url,
            "valid": cr.valid, "detail": cr.detail, "retries": cr.retries,
        })

    # ===== 完整版：整行标黄 + 汇总Sheet =====
    if output_full:
        for (sheet, row) in invalid_keys:
            ws = wb[sheet]
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col_idx).fill = FILL_YELLOW
        _build_summary_sheet(wb, result_list, valid_count, invalid_count, skipped_count)
        log.info(f"保存完整版: {output_path}")
        _save_workbook(wb, output_path)
    wb.close()

    # ===== 仅标黄版：整行标黄，不新增Sheet =====
    light_path = ""
    if output_light:
        light_path = _generate_light_output_path(input_path)
        log.info(f"保存仅标黄版: {light_path}")
        wb_light = load_workbook(input_path)
        for (sheet, row) in invalid_keys:
            ws = wb_light[sheet]
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col_idx).fill = FILL_YELLOW
        _save_workbook(wb_light, light_path)
        wb_light.close()

    log.info(f"检测完成: 有效={valid_count}, 无效={invalid_count}, 跳过={skipped_count}")

    return ProcessResult(
        input_path=input_path,
        output_path=output_path,
        light_output_path=light_path,
        total_urls=total_urls,
        valid_count=valid_count,
        invalid_count=invalid_count,
        skipped_count=skipped_count,
        results=result_list,
    )


def _save_workbook(wb, path: str) -> str:
    """保存工作簿，处理文件被占用的情况"""
    try:
        wb.save(path)
        return path
    except PermissionError:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        base, ext = os.path.splitext(path)
        alt = f"{base}_{ts}{ext}"
        log.warning(f"文件被占用，另存为: {alt}")
        wb.save(alt)
        return alt

def save_full_result(input_path: str, result: ProcessResult) -> str:
    """检测完成后，按需生成完整版输出（整行标黄 + 汇总Sheet）"""
    wb = load_workbook(input_path)
    for r in result.results:
        if not r["valid"]:
            ws = wb[r["sheet"]]
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=r["row"], column=col_idx).fill = FILL_YELLOW
    _build_summary_sheet(wb, result.results, result.valid_count, result.invalid_count, result.skipped_count)
    path = _save_workbook(wb, result.output_path)
    wb.close()
    log.info(f"完整版已保存: {path}")
    return path


def save_light_result(input_path: str, result: ProcessResult) -> str:
    """检测完成后，按需生成仅标黄版输出（整行标黄，无汇总Sheet）"""
    wb = load_workbook(input_path)
    for r in result.results:
        if not r["valid"]:
            ws = wb[r["sheet"]]
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=r["row"], column=col_idx).fill = FILL_YELLOW
    light_path = _generate_light_output_path(input_path)
    path = _save_workbook(wb, light_path)
    wb.close()
    log.info(f"仅标黄版已保存: {path}")
    return path


def _build_summary_sheet(wb, result_list, valid_count, invalid_count, skipped_count):
    """构建检测结果汇总Sheet"""
    sheet_name = "链接检测结果"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name, 0)

    # 统计行
    ws.cell(row=1, column=1, value=f"检测时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    ws.cell(row=2, column=1, value=f"总链接: {len(result_list)}  ✅ 有效: {valid_count}  ❌ 无效: {invalid_count}  ⚠ 跳过: {skipped_count}")
    ws.cell(row=2, column=1).font = Font(bold=True, size=11)

    # 表头
    headers = ["Sheet", "行号", "列名", "链接", "状态", "详情", "重试次数"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    for row_idx, entry in enumerate(result_list, 5):
        values = [
            entry["sheet"], entry["row"], entry["col"],
            entry["url"],
            "有效" if entry["valid"] else "无效",
            entry["detail"],
            entry["retries"],
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            if col_idx == 5:  # 状态列
                cell.fill = FILL_VALID if entry["valid"] else FILL_INVALID
                cell.font = FONT_VALID if entry["valid"] else FONT_INVALID
                cell.alignment = Alignment(horizontal="center")
            if col_idx == 4:  # 链接列
                cell.font = Font(color="0563C1", underline="single")

    col_widths = [18, 8, 18, 60, 10, 22, 10]
    for i, w in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w


# ==================== 工具函数 ====================
def generate_output_path(input_path: str) -> str:
    """生成完整版输出路径，输出到 result/ 子文件夹"""
    input_dir = os.path.dirname(input_path) or "."
    out_dir = os.path.join(input_dir, "result")
    os.makedirs(out_dir, exist_ok=True)
    base = os.path.splitext(os.path.basename(input_path))[0]
    ext = os.path.splitext(input_path)[1]
    ts = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    path = os.path.join(out_dir, f"{base}_链接检测结果_{ts}{ext}")
    if os.path.exists(path):
        counter = 1
        while os.path.exists(os.path.join(out_dir, f"{base}_链接检测结果_{ts}_{counter}{ext}")):
            counter += 1
        path = os.path.join(out_dir, f"{base}_链接检测结果_{ts}_{counter}{ext}")
    return path


def _generate_light_output_path(input_path: str) -> str:
    """生成仅标黄版输出路径，输出到 result/ 子文件夹"""
    input_dir = os.path.dirname(input_path) or "."
    out_dir = os.path.join(input_dir, "result")
    os.makedirs(out_dir, exist_ok=True)
    base = os.path.splitext(os.path.basename(input_path))[0]
    ext = os.path.splitext(input_path)[1]
    ts = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    path = os.path.join(out_dir, f"{base}_链接检测_仅标黄_{ts}{ext}")
    if os.path.exists(path):
        counter = 1
        while os.path.exists(os.path.join(out_dir, f"{base}_链接检测_仅标黄_{ts}_{counter}{ext}")):
            counter += 1
        path = os.path.join(out_dir, f"{base}_链接检测_仅标黄_{ts}_{counter}{ext}")
    return path

def hash_file(path: str, algorithm: str = "sha256") -> str:
    """计算文件哈希，用于校验"""
    h = hashlib.new(algorithm)
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()
